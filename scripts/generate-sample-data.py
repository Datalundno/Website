#!/usr/bin/env python3
"""Generate branded DataLund starter sample workbooks.

One simple portfolio story, shared suite field names, and a Start here sheet
so newcomers can bind Gantt, Resource Load, and Task List in minutes.

Outputs (Website repo):
  public/downloads/DatalundSuiteSample.xlsx
  public/downloads/GanttSampleData.xlsx
  public/downloads/ResourceLoadSampleData.xlsx
  public/downloads/TaskListSampleData.xlsx

Run from Website repo root:
  python3 scripts/generate-sample-data.py
"""

from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "public" / "downloads"

# Brand tokens (match datalund.no)
INK = "0F3D36"
ACCENT = "1A6B5C"
MINT = "2DD4BF"
PAPER = "E8F2EE"
WHITE = "FFFFFF"
MUTED = "3D5A54"
LINE = "C5D9D2"
RAG_GREEN = "1B7F5A"
RAG_AMBER = "B86E00"
RAG_RED = "B42318"
RAG_GREEN_BG = "D8F3E7"
RAG_AMBER_BG = "FCEFC7"
RAG_RED_BG = "F9D8D5"

FILL_INK = PatternFill("solid", fgColor=INK)
FILL_ACCENT = PatternFill("solid", fgColor=ACCENT)
FILL_PAPER = PatternFill("solid", fgColor=PAPER)
FILL_HEADER = PatternFill("solid", fgColor=ACCENT)
FILL_WHITE = PatternFill("solid", fgColor=WHITE)
FILL_ALT = PatternFill("solid", fgColor="F3FAF7")

FONT_BRAND = Font(name="Calibri", size=22, bold=True, color=WHITE)
FONT_TITLE = Font(name="Calibri", size=14, bold=True, color=INK)
FONT_BODY = Font(name="Calibri", size=11, color=INK)
FONT_MUTED = Font(name="Calibri", size=11, color=MUTED)
FONT_STEP = Font(name="Calibri", size=11, bold=True, color=ACCENT)
FONT_HEADER = Font(name="Calibri", size=11, bold=True, color=WHITE)
FONT_LINK = Font(name="Calibri", size=11, color=ACCENT, underline="single")
FONT_SMALL = Font(name="Calibri", size=10, color=MUTED)

THIN = Border(
    left=Side(style="thin", color=LINE),
    right=Side(style="thin", color=LINE),
    top=Side(style="thin", color=LINE),
    bottom=Side(style="thin", color=LINE),
)

PROJECT_HEADERS = [
    "Project",
    "RAG",
    "Group",
    "Project lead",
    "Progress",
    "Start Date",
    "End Date",
    "Next milestone",
    "Obstacles",
    "Notes",
]

TASK_HEADERS = [
    "Task",
    "Start Date",
    "End Date",
    "Duration",
    "Progress",
    "Group",
    "Resource",
    "Project",
    "RAG",
]

# Compact starter portfolio — enough for RAG, overlaps, milestone, duration-only.
# Progress is percent 0–100. Dates sit around Aug 2026 for “time window from today”.
PROJECTS = [
    # Project, RAG, Group, Project lead, Progress, Start, End, milestone, obstacles, notes
    ("North Sea Hub", "Green", "Delivery", "Ada Ng", 72, date(2026, 6, 11), date(2026, 9, 19), "FAT complete", "", "On track for Q3"),
    ("Arctic Link", "Amber", "Delivery", "Bo Berg", 45, date(2026, 5, 12), date(2026, 8, 30), "Cable pull", "Weather window", "Watch supplier lead time"),
    ("Yard Retrofit", "Red", "Initiation", "Cara Diaz", 18, date(2026, 7, 11), date(2026, 8, 20), "Scope freeze", "Budget overrun", "Critical path slip"),
    ("Digital Twin v2", "Green", "Build", "Dan Okonkwo", 88, date(2026, 4, 12), date(2026, 8, 25), "UAT", "", "Near release"),
    ("Port Expansion", "Amber", "Plan", "Finn Olsen", 12, date(2026, 8, 15), date(2027, 2, 6), "Permits", "", "Replace these rows with your projects"),
]

TASKS = [
    # Task, Start, End, Duration, Progress, Group, Resource, Project, RAG
    ("Requirements", date(2026, 6, 11), date(2026, 6, 25), None, 100, "Delivery", "Ada Ng", "North Sea Hub", "Green"),
    ("Wireframes", date(2026, 6, 20), date(2026, 7, 5), None, 90, "Delivery", "Jordan Lee", "North Sea Hub", "Green"),
    ("API design", date(2026, 7, 1), date(2026, 7, 20), None, 75, "Delivery", "Sam Ortiz", "North Sea Hub", "Green"),
    ("FAT prep", date(2026, 8, 20), date(2026, 9, 5), None, 20, "Delivery", "Ada Ng", "North Sea Hub", "Amber"),
    ("Cable survey", date(2026, 5, 12), date(2026, 6, 15), None, 100, "Delivery", "Bo Berg", "Arctic Link", "Green"),
    ("Cable pull", date(2026, 7, 1), date(2026, 8, 20), None, 40, "Delivery", "Bo Berg", "Arctic Link", "Amber"),
    ("Weather hold", date(2026, 8, 10), date(2026, 8, 18), None, 0, "Delivery", "Riley Chen", "Arctic Link", "Red"),
    ("On-call week", date(2026, 8, 18), date(2026, 8, 24), None, 0, "Delivery", "Bo Berg", "Arctic Link", "Red"),
    ("Scope workshop", date(2026, 7, 11), date(2026, 7, 25), None, 50, "Initiation", "Cara Diaz", "Yard Retrofit", "Amber"),
    ("Budget review", date(2026, 7, 20), date(2026, 8, 5), None, 15, "Initiation", "Finn Olsen", "Yard Retrofit", "Red"),
    ("Scope freeze", date(2026, 8, 12), date(2026, 8, 12), None, 0, "Initiation", "Cara Diaz", "Yard Retrofit", "Red"),
    ("Model training", date(2026, 4, 12), date(2026, 6, 30), None, 100, "Build", "Dan Okonkwo", "Digital Twin v2", "Green"),
    ("Integration", date(2026, 7, 1), date(2026, 8, 10), None, 85, "Build", "Sam Ortiz", "Digital Twin v2", "Green"),
    ("UAT", date(2026, 8, 11), date(2026, 8, 25), None, 30, "Build", "Ada Ng", "Digital Twin v2", "Amber"),
    ("QA soak", date(2026, 8, 20), None, 14.0, 10, "Build", "Riley Chen", "Digital Twin v2", "Green"),
    ("Permit draft", date(2026, 8, 15), date(2026, 10, 1), None, 10, "Plan", "Finn Olsen", "Port Expansion", "Amber"),
    ("Stakeholder map", date(2026, 8, 20), date(2026, 9, 15), None, 5, "Plan", "Jordan Lee", "Port Expansion", "Green"),
]


def set_col_widths(ws: Worksheet, widths: list[float]) -> None:
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def paint_banner(ws: Worksheet, title: str, subtitle: str, cols: int = 6) -> None:
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=cols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=cols)
    c1 = ws.cell(1, 1, title)
    c1.font = FONT_BRAND
    c1.fill = FILL_INK
    c1.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[1].height = 36
    c2 = ws.cell(2, 1, subtitle)
    c2.font = Font(name="Calibri", size=11, color=INK)
    c2.fill = FILL_PAPER
    c2.alignment = Alignment(vertical="center", horizontal="left", indent=1)
    ws.row_dimensions[2].height = 22
    for col in range(1, cols + 1):
        ws.cell(1, col).fill = FILL_INK
        ws.cell(2, col).fill = FILL_PAPER


def write_start_block(ws: Worksheet, lines: list[tuple[str, str]]) -> int:
    """Write (style, text) rows starting at row 4. Styles: title|step|body|muted|blank|link."""
    row = 4
    for style, text in lines:
        cell = ws.cell(row, 1, text if text else None)
        if style == "title":
            cell.font = FONT_TITLE
            ws.row_dimensions[row].height = 20
        elif style == "step":
            cell.font = FONT_STEP
        elif style == "muted":
            cell.font = FONT_MUTED
        elif style == "link":
            cell.font = FONT_LINK
            if text.startswith("http"):
                cell.hyperlink = text
        elif style == "blank":
            cell.value = None
        else:
            cell.font = FONT_BODY
        row += 1
    return row


def style_table(ws: Worksheet, headers: list[str], n_rows: int, date_cols: set[int], progress_cols: set[int], rag_cols: set[int]) -> None:
    header_row = 1
    for col, name in enumerate(headers, start=1):
        cell = ws.cell(header_row, col, name)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = THIN
    ws.row_dimensions[header_row].height = 20
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{n_rows + 1}"
    ws.freeze_panes = "A2"

    for r in range(2, n_rows + 2):
        for c in range(1, len(headers) + 1):
            cell = ws.cell(r, c)
            cell.border = THIN
            cell.font = FONT_BODY
            if r % 2 == 0:
                cell.fill = FILL_ALT
            if c in date_cols and cell.value is not None:
                cell.number_format = "YYYY-MM-DD"
            if c in progress_cols and isinstance(cell.value, (int, float)):
                cell.number_format = '0"%"'
                cell.alignment = Alignment(horizontal="right")
            if c in rag_cols and isinstance(cell.value, str):
                rag = cell.value.strip().lower()
                if rag in {"green", "grønn", "gronn"}:
                    cell.fill = PatternFill("solid", fgColor=RAG_GREEN_BG)
                    cell.font = Font(name="Calibri", size=11, bold=True, color=RAG_GREEN)
                elif rag in {"amber", "gul", "yellow"}:
                    cell.fill = PatternFill("solid", fgColor=RAG_AMBER_BG)
                    cell.font = Font(name="Calibri", size=11, bold=True, color=RAG_AMBER)
                elif rag in {"red", "rød", "rod"}:
                    cell.fill = PatternFill("solid", fgColor=RAG_RED_BG)
                    cell.font = Font(name="Calibri", size=11, bold=True, color=RAG_RED)


def write_data_sheet(ws: Worksheet, headers: list[str], rows: list[tuple], widths: list[float], date_names: set[str], progress_names: set[str], rag_names: set[str]) -> None:
    ws.append(headers)
    for row in rows:
        ws.append(list(row))
    date_cols = {i for i, h in enumerate(headers, start=1) if h in date_names}
    progress_cols = {i for i, h in enumerate(headers, start=1) if h in progress_names}
    rag_cols = {i for i, h in enumerate(headers, start=1) if h in rag_names}
    style_table(ws, headers, len(rows), date_cols, progress_cols, rag_cols)
    set_col_widths(ws, widths)


def add_data_sheet(wb: Workbook, title: str, headers: list[str], rows: list[tuple], widths: list[float], tab_color: str = ACCENT) -> Worksheet:
    ws = wb.create_sheet(title)
    ws.sheet_properties.tabColor = tab_color
    write_data_sheet(
        ws,
        headers,
        rows,
        widths,
        date_names={"Start Date", "End Date"},
        progress_names={"Progress"},
        rag_names={"RAG"},
    )
    return ws


def build_suite_start(ws: Worksheet) -> None:
    ws.sheet_properties.tabColor = INK
    paint_banner(
        ws,
        "DataLund suite sample",
        "A simple starter workbook for Task List, Gantt, and Resource Load",
        cols=6,
    )
    write_start_block(
        ws,
        [
            ("title", "Quick start"),
            ("step", "1. Import the visuals"),
            ("body", "Download the suite zip from datalund.no, then in Power BI Desktop:"),
            ("body", "Visualizations … → Import a visual from a file → pick each .pbiviz."),
            ("blank", ""),
            ("step", "2. Load this workbook"),
            ("body", "Home → Get data → Excel workbook → load Projects and Tasks."),
            ("body", "Relate Tasks[Project] → Projects[Project] (many-to-one)."),
            ("blank", ""),
            ("step", "3. Bind the same columns everywhere"),
            ("body", "Minimum for Task List: Project (+ RAG and Progress look great)."),
            ("body", "Minimum for Gantt / Resource Load: Task, Start Date, and End Date or Duration."),
            ("body", "Then add Progress, Group, and Resource / Project lead."),
            ("blank", ""),
            ("step", "4. Build one page"),
            ("body", "Put Task List on the left. Click a project — Gantt and Resource Load follow."),
            ("blank", ""),
            ("title", "What’s in the sheets"),
            ("body", "Projects — portfolio rows for DataLund Task List (Lists-style names)."),
            ("body", "Tasks — assignment rows for DataLund Gantt and Resource Load."),
            ("muted", "Progress is 0–100. RAG is Red / Amber / Green (also Rød / Gul / Grønn)."),
            ("muted", "Scope freeze is a milestone (same start and end). QA soak uses Duration instead of End Date."),
            ("blank", ""),
            ("title", "Make it yours"),
            ("body", "Replace the sample rows with your projects and tasks. Keep the column headers."),
            ("body", "Same field names work across the whole DataLund suite."),
            ("blank", ""),
            ("link", "https://datalund.no"),
            ("muted", "Free Power BI visuals · support@datalund.no"),
        ],
    )
    set_col_widths(ws, [92, 14, 14, 14, 14, 14])
    ws.row_dimensions[4].height = 22


def build_visual_start(ws: Worksheet, visual: str, subtitle: str, steps: list[tuple[str, str]]) -> None:
    ws.sheet_properties.tabColor = INK
    paint_banner(ws, f"DataLund {visual}", subtitle, cols=5)
    write_start_block(ws, steps)
    set_col_widths(ws, [88, 14, 14, 14, 14])


def build_suite() -> Path:
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Start here"
    build_suite_start(ws0)

    add_data_sheet(
        wb,
        "Projects",
        PROJECT_HEADERS,
        PROJECTS,
        widths=[18, 10, 12, 14, 10, 12, 12, 16, 18, 28],
        tab_color=MINT,
    )
    add_data_sheet(
        wb,
        "Tasks",
        TASK_HEADERS,
        TASKS,
        widths=[16, 12, 12, 10, 10, 12, 14, 16, 10],
        tab_color=ACCENT,
    )

    # Put Start here first (already is), remove default ordering issues
    path = OUT / "DatalundSuiteSample.xlsx"
    wb.save(path)
    return path


def build_gantt() -> Path:
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Start here"
    build_visual_start(
        ws0,
        "Gantt chart",
        "Starter tasks for the timeline visual",
        [
            ("title", "Quick start"),
            ("step", "1. Import ganttChart.pbiviz"),
            ("body", "Visualizations … → Import a visual from a file."),
            ("step", "2. Load the Tasks sheet from this workbook"),
            ("step", "3. Bind"),
            ("body", "Task → Task · Start Date → Start Date · End Date → End Date"),
            ("body", "Optional: Duration, Progress, Group, Resource"),
            ("blank", ""),
            ("muted", "Same column names as Resource Load and Task List."),
            ("muted", "Want the full suite page? Use DatalundSuiteSample.xlsx."),
            ("blank", ""),
            ("link", "https://datalund.no/visuals/gantt/"),
        ],
    )
    rows = [(t[0], t[1], t[2], t[3], t[4], t[5], t[6]) for t in TASKS]
    add_data_sheet(
        wb,
        "Tasks",
        ["Task", "Start Date", "End Date", "Duration", "Progress", "Group", "Resource"],
        rows,
        widths=[16, 12, 12, 10, 10, 12, 14],
    )
    path = OUT / "GanttSampleData.xlsx"
    wb.save(path)
    return path


def build_resource_load() -> Path:
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Start here"
    build_visual_start(
        ws0,
        "Resource Load",
        "Starter assignments for the people timeline",
        [
            ("title", "Quick start"),
            ("step", "1. Import resourceLoad.pbiviz"),
            ("body", "Visualizations … → Import a visual from a file."),
            ("step", "2. Load the Assignments sheet from this workbook"),
            ("step", "3. Bind"),
            ("body", "Resource → Resource · Task → Task · Start Date → Start Date"),
            ("body", "End Date → End Date · Duration → Duration (optional: Progress, Group)"),
            ("muted", "Put RAG on Tooltips if you want status on hover."),
            ("blank", ""),
            ("muted", "Same column names as Gantt and Task List."),
            ("muted", "Want the full suite page? Use DatalundSuiteSample.xlsx."),
            ("blank", ""),
            ("link", "https://datalund.no/visuals/resource-load/"),
        ],
    )
    rows = [(t[6], t[0], t[1], t[2], t[3], t[4], t[5], t[8]) for t in TASKS]
    add_data_sheet(
        wb,
        "Assignments",
        ["Resource", "Task", "Start Date", "End Date", "Duration", "Progress", "Group", "RAG"],
        rows,
        widths=[14, 16, 12, 12, 10, 10, 12, 10],
    )
    path = OUT / "ResourceLoadSampleData.xlsx"
    wb.save(path)
    return path


def build_task_list() -> Path:
    wb = Workbook()
    ws0 = wb.active
    ws0.title = "Start here"
    build_visual_start(
        ws0,
        "Task List",
        "Starter projects for the portfolio list",
        [
            ("title", "Quick start"),
            ("step", "1. Import taskList.pbiviz"),
            ("body", "Visualizations … → Import a visual from a file."),
            ("step", "2. Load the Projects sheet from this workbook"),
            ("step", "3. Bind"),
            ("body", "Project → Project · RAG → RAG · Group → Group · Project lead → Project lead"),
            ("body", "Progress → Progress · Start Date → Start Date · End Date → End Date"),
            ("body", "Tooltips → Next milestone, Obstacles, Notes"),
            ("blank", ""),
            ("muted", "Project = Task well · Project lead = Resource well · RAG = status."),
            ("muted", "Want Gantt + Resource Load on the same page? Use DatalundSuiteSample.xlsx."),
            ("blank", ""),
            ("link", "https://datalund.no/visuals/task-list/"),
        ],
    )
    add_data_sheet(
        wb,
        "Projects",
        PROJECT_HEADERS,
        PROJECTS,
        widths=[18, 10, 12, 14, 10, 12, 12, 16, 18, 28],
        tab_color=MINT,
    )
    path = OUT / "TaskListSampleData.xlsx"
    wb.save(path)
    return path


def main() -> None:
    OUT.mkdir(parents=True, exist_ok=True)
    paths = [build_suite(), build_gantt(), build_resource_load(), build_task_list()]
    for p in paths:
        print(f"wrote {p.relative_to(ROOT)} ({p.stat().st_size} bytes)")


if __name__ == "__main__":
    main()
